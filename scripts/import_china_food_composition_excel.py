from __future__ import annotations

import argparse
import re
import sqlite3
from dataclasses import dataclass
from pathlib import Path

from openpyxl import load_workbook


PROJECT_ROOT = Path(__file__).resolve().parents[1]
SCHEMA_PATH = PROJECT_ROOT / "knowledge_base" / "structured" / "schema.sql"
DEFAULT_DB_PATH = PROJECT_ROOT / "data" / "clinical_foods.db"
DEFAULT_XLSX_PATH = (
    PROJECT_ROOT
    / "data"
    / "raw"
    / "china_food_composition"
    / "china_food_composition.xlsx"
)

SOURCE_TITLE = "用户提供：各种食物营养成分表.xlsx"
SOURCE_VERSION = "Excel import 2026-04-25"


@dataclass(frozen=True)
class ImportStats:
    foods: int
    nutrients: int
    aliases: int
    allergens: int
    risk_tags: int


def main() -> None:
    args = parse_args()
    db_path = Path(args.db).resolve()
    xlsx_path = Path(args.xlsx).resolve()
    if not xlsx_path.exists():
        raise FileNotFoundError(f"Excel file not found: {xlsx_path}")

    db_path.parent.mkdir(parents=True, exist_ok=True)
    with sqlite3.connect(db_path) as conn:
        conn.execute("PRAGMA foreign_keys = ON")
        conn.executescript(SCHEMA_PATH.read_text(encoding="utf-8"))
        ensure_schema_migrations(conn)
        stats = import_excel(conn, xlsx_path, replace_existing=not args.append)
        conn.commit()

    print(
        "China food composition import complete: "
        f"db={db_path}, foods={stats.foods}, nutrients={stats.nutrients}, "
        f"aliases={stats.aliases}, allergens={stats.allergens}, risk_tags={stats.risk_tags}"
    )


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Import a Chinese food composition Excel into clinical_foods.db")
    parser.add_argument("--xlsx", default=str(DEFAULT_XLSX_PATH), help="Input Excel file path.")
    parser.add_argument("--db", default=str(DEFAULT_DB_PATH), help="SQLite database path.")
    parser.add_argument("--append", action="store_true", help="Append/update without clearing prior rows from this source.")
    return parser.parse_args()


def ensure_schema_migrations(conn: sqlite3.Connection) -> None:
    food_columns = {row[1] for row in conn.execute("PRAGMA table_info(food_items)").fetchall()}
    if "source_food_id" not in food_columns:
        conn.execute("ALTER TABLE food_items ADD COLUMN source_food_id TEXT")

    nutrient_columns = {row[1] for row in conn.execute("PRAGMA table_info(food_nutrients_per_100g)").fetchall()}
    for column, sql_type in EXTRA_NUTRIENT_COLUMNS.items():
        if column not in nutrient_columns:
            conn.execute(f"ALTER TABLE food_nutrients_per_100g ADD COLUMN {column} {sql_type}")
    conn.execute(
        "CREATE INDEX IF NOT EXISTS idx_food_items_source_food_id ON food_items(source_food_id)"
    )


def import_excel(conn: sqlite3.Connection, xlsx_path: Path, replace_existing: bool = True) -> ImportStats:
    source_id = upsert_source(conn, xlsx_path)
    if replace_existing:
        conn.execute("DELETE FROM food_items WHERE source_id = ?", (source_id,))

    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb.active
    header = [normalize_header(value) for value in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]
    index = {name: idx for idx, name in enumerate(header)}
    required = ["食物", "能量", "蛋白质", "糖类", "脂肪"]
    missing = [name for name in required if name not in index]
    if missing:
        raise ValueError(f"Excel is missing required columns: {missing}")

    stats = {"foods": 0, "nutrients": 0, "aliases": 0, "allergens": 0, "risk_tags": 0}
    for row_number, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        food_name = clean_text(row[index["食物"]] if index["食物"] < len(row) else "")
        if not food_name:
            continue

        nutrients = extract_nutrients(row, index)
        food_id = upsert_food(conn, source_id, food_name, row_number)
        stats["foods"] += 1
        if upsert_nutrients(conn, food_id, source_id, nutrients):
            stats["nutrients"] += 1
        stats["aliases"] += upsert_aliases(conn, food_id, food_name)
        upsert_allergen_flags(conn, food_id, source_id, food_name)
        stats["allergens"] += 1
        stats["risk_tags"] += upsert_risk_tags(conn, food_id, source_id, food_name, nutrients)

    return ImportStats(**stats)


def upsert_source(conn: sqlite3.Connection, xlsx_path: Path) -> int:
    row = conn.execute(
        "SELECT source_id FROM source_documents WHERE source_title = ? AND version = ?",
        (SOURCE_TITLE, SOURCE_VERSION),
    ).fetchone()
    if row:
        return int(row[0])
    cursor = conn.execute(
        """
        INSERT INTO source_documents (
            source_title, source_org, publish_year, version, source_url,
            evidence_level, license_note, last_reviewed_at
        )
        VALUES (?, ?, ?, ?, ?, ?, ?, ?)
        """,
        (
            SOURCE_TITLE,
            "user_provided",
            2026,
            SOURCE_VERSION,
            str(xlsx_path),
            "user_provided_food_composition_table",
            "User-provided Excel. Use locally/internal only unless publication and redistribution rights are confirmed.",
            "2026-04-25",
        ),
    )
    return int(cursor.lastrowid)


def upsert_food(conn: sqlite3.Connection, source_id: int, food_name: str, row_number: int) -> int:
    row = conn.execute(
        "SELECT food_id FROM food_items WHERE canonical_name = ?",
        (food_name,),
    ).fetchone()
    if row:
        food_id = int(row[0])
        conn.execute(
            """
            UPDATE food_items
            SET chinese_name = ?, food_category = ?, processing_level = ?, source_food_id = ?,
                notes = ?, source_id = ?, updated_at = CURRENT_TIMESTAMP
            WHERE food_id = ?
            """,
            (
                food_name,
                infer_food_category(food_name),
                infer_processing_level(food_name),
                f"china_excel_row_{row_number}",
                f"Imported from user-provided Excel row {row_number}",
                source_id,
                food_id,
            ),
        )
        return food_id

    cursor = conn.execute(
        """
        INSERT INTO food_items (
            canonical_name, chinese_name, food_category, processing_level,
            source_food_id, notes, source_id
        )
        VALUES (?, ?, ?, ?, ?, ?, ?)
        """,
        (
            food_name,
            food_name,
            infer_food_category(food_name),
            infer_processing_level(food_name),
            f"china_excel_row_{row_number}",
            f"Imported from user-provided Excel row {row_number}",
            source_id,
        ),
    )
    return int(cursor.lastrowid)


def extract_nutrients(row: tuple, index: dict[str, int]) -> dict[str, float | None]:
    mapping = {
        "energy_kcal": "能量",
        "protein_g": "蛋白质",
        "carbohydrate_g": "糖类",
        "fat_g": "脂肪",
        "dietary_fiber_g": "纤维",
        "potassium_mg": "钾",
        "sodium_mg": "钠",
        "calcium_mg": "钙",
        "magnesium_mg": "镁",
        "iron_mg": "铁",
        "manganese_mg": "锰",
        "zinc_mg": "锌",
        "copper_mg": "铜",
        "phosphorus_mg": "磷",
        "selenium_ug": "硒",
        "vitamin_a_ug": "VA",
        "beta_carotene_ug": "胡萝卜素",
        "retinol_equivalent_ug": "视黄醇当量",
        "thiamin_mg": "VB1",
        "riboflavin_mg": "VB2",
        "niacin_mg": "烟酸",
        "vitamin_c_mg": "VC",
        "vitamin_e_mg": "VE",
    }
    nutrients: dict[str, float | None] = {}
    for target, source in mapping.items():
        col = index.get(source)
        nutrients[target] = parse_float(row[col]) if col is not None and col < len(row) else None
    return nutrients


def upsert_nutrients(conn: sqlite3.Connection, food_id: int, source_id: int, nutrients: dict[str, float | None]) -> bool:
    if not any(value is not None for value in nutrients.values()):
        return False
    columns = list(nutrients)
    conn.execute(
        f"""
        INSERT INTO food_nutrients_per_100g (
            food_id, {", ".join(columns)}, data_quality, source_id
        )
        VALUES (?, {", ".join("?" for _ in columns)}, ?, ?)
        ON CONFLICT(food_id, source_id) DO UPDATE SET
            {", ".join(f"{column} = excluded.{column}" for column in columns)},
            data_quality = excluded.data_quality,
            updated_at = CURRENT_TIMESTAMP
        """,
        [food_id, *[nutrients[column] for column in columns], "china_excel_user_provided", source_id],
    )
    return True


def upsert_aliases(conn: sqlite3.Connection, food_id: int, food_name: str) -> int:
    aliases = {food_name, normalize_food_alias(food_name)}
    base = re.sub(r"[（(].*?[）)]", "", food_name).strip()
    if base:
        aliases.add(base)
    if food_name == "牛乳":
        aliases.add("牛奶")
    if "酸牛乳" in food_name or food_name == "牛乳(酸)":
        aliases.add("酸奶")
    if food_name.startswith("米饭"):
        aliases.add("米饭")
    if food_name.startswith("稻米"):
        aliases.add("大米")
    inserted = 0
    for alias in sorted(alias for alias in aliases if alias):
        before = conn.total_changes
        conn.execute(
            "INSERT OR IGNORE INTO food_aliases (food_id, alias, language) VALUES (?, ?, ?)",
            (food_id, alias, "zh"),
        )
        inserted += conn.total_changes - before
    return inserted


def upsert_allergen_flags(conn: sqlite3.Connection, food_id: int, source_id: int, food_name: str) -> None:
    flags = infer_allergens(food_name)
    conn.execute(
        """
        INSERT INTO allergen_flags (
            food_id, contains_gluten, contains_peanut, contains_tree_nut,
            contains_crustacean, contains_soy, contains_dairy, contains_egg,
            cross_contamination_risk, source_id
        )
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        ON CONFLICT(food_id) DO UPDATE SET
            contains_gluten = excluded.contains_gluten,
            contains_peanut = excluded.contains_peanut,
            contains_tree_nut = excluded.contains_tree_nut,
            contains_crustacean = excluded.contains_crustacean,
            contains_soy = excluded.contains_soy,
            contains_dairy = excluded.contains_dairy,
            contains_egg = excluded.contains_egg,
            cross_contamination_risk = excluded.cross_contamination_risk,
            source_id = excluded.source_id,
            updated_at = CURRENT_TIMESTAMP
        """,
        (
            food_id,
            flags["contains_gluten"],
            flags["contains_peanut"],
            flags["contains_tree_nut"],
            flags["contains_crustacean"],
            flags["contains_soy"],
            flags["contains_dairy"],
            flags["contains_egg"],
            "Heuristic from Chinese food name; verify ingredient labels for packaged foods.",
            source_id,
        ),
    )


def upsert_risk_tags(
    conn: sqlite3.Connection,
    food_id: int,
    source_id: int,
    food_name: str,
    nutrients: dict[str, float | None],
) -> int:
    tags: list[tuple[str, str, str]] = []
    if any(token in food_name for token in ["西柚", "葡萄柚"]):
        tags.append(("cyp3a4_interaction_food", "drug_food_interaction", "西柚/葡萄柚可与部分药物发生相互作用。"))
    if any(token in food_name for token in ["酒", "啤酒", "白酒", "黄酒", "葡萄酒"]):
        tags.append(("alcohol", "diabetes", "酒精可能增加低血糖和能量摄入风险。"))
    if (nutrients.get("sodium_mg") or 0) >= 400:
        tags.append(("high_sodium", "hypertension", "钠含量 >= 400 mg/100g。"))
    if (nutrients.get("potassium_mg") or 0) >= 300:
        tags.append(("high_potassium", "chronic_kidney_disease", "钾含量 >= 300 mg/100g。"))
    if (nutrients.get("phosphorus_mg") or 0) >= 250:
        tags.append(("high_phosphorus", "chronic_kidney_disease", "磷含量 >= 250 mg/100g。"))

    inserted = 0
    for risk_tag, condition, rationale in tags:
        before = conn.total_changes
        conn.execute(
            """
            INSERT OR IGNORE INTO food_risk_tags (
                food_id, risk_tag, applies_to_condition, rationale, source_id
            )
            VALUES (?, ?, ?, ?, ?)
            """,
            (food_id, risk_tag, condition, rationale, source_id),
        )
        inserted += conn.total_changes - before
    return inserted


def infer_food_category(food_name: str) -> str:
    category_rules = [
        ("alcoholic_beverage", ["酒", "啤", "葡萄酒"]),
        ("egg", ["蛋"]),
        ("dairy", ["奶", "乳酪", "奶酪", "酸奶"]),
        ("dairy", ["牛乳", "羊乳", "炼乳", "乳粉"]),
        ("soybean_product", ["豆腐", "豆浆", "豆干", "腐竹", "黄豆", "大豆"]),
        ("grain", ["米", "面", "粉", "馒头", "包子", "饼", "粥", "麦", "玉米"]),
        ("vegetable", ["菜", "瓜", "笋", "菇", "菌", "藕", "萝卜", "番茄", "茄子", "菠菜"]),
        ("fruit", ["苹果", "梨", "桃", "橙", "柑", "香蕉", "葡萄", "西瓜", "柚", "莓", "枣"]),
        ("meat", ["猪", "牛", "羊", "鸡", "鸭", "鹅", "肉", "肝", "肾", "心"]),
        ("aquatic", ["鱼", "虾", "蟹", "贝", "蛤", "鳝", "鳟", "鲤", "鲫", "鲈"]),
        ("nut_seed", ["花生", "芝麻", "杏仁", "核桃", "瓜子", "松子", "榛子"]),
        ("oil", ["油"]),
    ]
    for category, tokens in category_rules:
        if any(token in food_name for token in tokens):
            return category
    return "china_food_composition"


def infer_processing_level(food_name: str) -> str:
    if any(token in food_name for token in ["熟", "罐头", "腌", "酱", "糕", "饼", "酒", "油", "糖"]):
        return "processed"
    return "unspecified"


def infer_allergens(food_name: str) -> dict[str, int]:
    return {
        "contains_gluten": int(any(token in food_name for token in ["小麦", "面粉", "面包", "馒头", "面条", "饼干", "包子"])),
        "contains_peanut": int("花生" in food_name),
        "contains_tree_nut": int(any(token in food_name for token in ["杏仁", "核桃", "腰果", "榛子", "松子", "开心果"])),
        "contains_crustacean": int(any(token in food_name for token in ["虾", "蟹", "龙虾"])),
        "contains_soy": int(any(token in food_name for token in ["黄豆", "大豆", "豆腐", "豆浆", "豆干", "腐竹"])),
        "contains_dairy": int(any(token in food_name for token in ["牛奶", "羊奶", "酸奶", "奶酪", "乳酪", "奶粉"])),
        "contains_egg": int("蛋" in food_name),
    }


def normalize_header(value) -> str:
    text = clean_text(value).replace("（", "(").replace("）", ")")
    text = re.sub(r"\(.*?\)", "", text)
    return text.strip()


def normalize_food_alias(value: str) -> str:
    return clean_text(value).replace("（", "(").replace("）", ")")


def clean_text(value) -> str:
    return " ".join(str(value or "").replace("\ufeff", "").split()).strip()


def parse_float(value) -> float | None:
    if value is None or value == "":
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


EXTRA_NUTRIENT_COLUMNS = {
    "magnesium_mg": "REAL",
    "iron_mg": "REAL",
    "manganese_mg": "REAL",
    "zinc_mg": "REAL",
    "copper_mg": "REAL",
    "selenium_ug": "REAL",
    "vitamin_a_ug": "REAL",
    "beta_carotene_ug": "REAL",
    "retinol_equivalent_ug": "REAL",
    "thiamin_mg": "REAL",
    "riboflavin_mg": "REAL",
    "niacin_mg": "REAL",
    "vitamin_c_mg": "REAL",
    "vitamin_e_mg": "REAL",
}


if __name__ == "__main__":
    main()
